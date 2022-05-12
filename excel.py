# -----------------------------------------
#    _____  _____ ___| |  _ __  _   _
#   / _ \ \/ / __/ _ \ | | '_ \| | | |
#  |  __/>  < (_|  __/ |_| |_) | |_| |
#   \___/_/\_\___\___|_(_) .__/ \__, |
#                         |_|    |___/
# -----------------------------------------
"""
this file is used to read and write excel file

"""

import openpyxl
import xlrd
import xlwt
from .__log import *
from .__base import file_suffix

EXCEL_LOG = Log()


def read_xls_row(path, sheet_name=None):
    workbook = xlrd.open_workbook(path)
    if sheet_name is None:
        sheet_names = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheet_names[0])
    else:
        worksheet = workbook.sheet_by_name(sheet_name)
    data = []
    for i in range(0, worksheet.nrows):
        tmp_list = []
        for j in range(0, worksheet.ncols):
            tmp_list.append(worksheet.cell_value(i, j))
        data.append(tmp_list)
    return data


def read_xls_column(path, sheet_name=None):
    """
    read the xls file by column 
    """
    workbook = xlrd.open_workbook(path)
    if sheet_name is None:
        sheet_names = workbook.sheet_names()
        worksheet = workbook.sheet_by_name(sheet_names[0])
    else:
        worksheet = workbook.sheet_by_name(sheet_name)
    data = {}
    for i in range(worksheet.ncols):
        tmp_list = []
        for j in range(worksheet.nrows):
            tmp_list.append(worksheet.cell_value(j, i))
        data_key = str(tmp_list.pop(0))
        data[data_key] = tmp_list
    return data


def write_xls_excel(path, sheet_name, two_dimensional_data):
    '''
    :param path:
    :param sheet_name:
    :param two_dimensional_data:
    :return:
    '''
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet(sheet_name)
    for i in range(0, len(two_dimensional_data)):
        for j in range(0, len(two_dimensional_data[i])):
            sheet.write(i, j, two_dimensional_data[i][j])
    workbook.save(path)
    EXCEL_LOG.complete('Save the data success!')


def read_xlsx_row(path, sheet_name=None):
    """
    读取xlsx格式文件
    参数：
        path:文件路径
        sheet_name:表名
    返回：
        data:表格中的数据
    """
    workbook = openpyxl.load_workbook(path, read_only=True)
    if sheet_name is None:
        sheet_names = workbook.sheetnames # 获取所有表名
        sheet = workbook[sheet_names[0]]
    else:
        sheet = workbook[sheet_name]
    data = []
    for row in sheet.rows:
        tmp_data = []
        for cell in row:
            tmp_data.append(cell.value)
        data.append(tmp_data)
    return data


def read_xlsx_column(path, sheet_name=None):
    workbook = openpyxl.load_workbook(path, read_only=True)
    if sheet_name is None:
        sheet_names = workbook.sheetnames
        sheet = workbook[sheet_names[0]]
    else:
        sheet = workbook[sheet_name]
    col_num = sheet.max_column
    data = {}
    rows = sheet.iter_rows(min_row=1, max_row=1)
    first_row = next(rows)
    headings = [cell.value for cell in first_row]
    if len(headings) != col_num:
        EXCEL_LOG.error(
            "The number of columns is not equal to the number of headings!")
        pass
    else:
        for i in range(col_num):
            tmp_list = [each_row[i].value for each_row in sheet.rows]
            del tmp_list[0]
            data[headings[i]] = tmp_list
    return data


def write_xlsx(path, two_dimensional_data, sheet_name='Sheet1'):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, len(two_dimensional_data)):
        for j in range(0, len(two_dimensional_data[i])):
            sheet.cell(row=i + 1,
                       column=j + 1,
                       value=str(two_dimensional_data[i][j]))
    workbook.save(path)
    EXCEL_LOG.complete('Save the data success!')


'''
追加写入xlsx格式文件
参数：
    path:文件路径
    sheet_name:表名
    two_dimensional_data：将要写入表格的数据（二维列表）
'''


def xlsx_append(path, two_dimensional_data, sheet_name='sheet1'):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for tdd in two_dimensional_data:
        sheet.append(tdd)
    workbook.save(path)
    EXCEL_LOG.complete('Append and write the data success!')


def excel_merge(dst_path, src_path):
    buffer_zone = read_xlsx_row(src_path)
    append_list = buffer_zone[1:]
    xlsx_append(src_path, append_list)
    pass


#TODO <++> 待完善
def save_spider(spider: list, path='spider.xlsx', sheet_name="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if isinstance(spider[0], dict):
        key_list = list(spider[0].keys())
        ws.append(key_list)
        for item in spider:
            value_list = list(item.values())
            ws.append(value_list)
        wb.save(path)
    else:
        EXCEL_LOG.error(
            'Maybe you can use write_xlsx function to save your data!')


def read_row(path, sheet_name=None):
    if file_suffix(path) == 'xls':
        return read_xls_row(path, sheet_name)
    elif file_suffix(path) == 'xlsx':
        return read_xlsx_row(path, sheet_name)
    else:
        EXCEL_LOG.error('The file format is not supported!')
        pass


def read_column(path, sheet_name=None):
    if file_suffix(path) == 'xls':
        return read_xls_column(path, sheet_name)
    elif file_suffix(path) == 'xlsx':
        return read_xlsx_column(path, sheet_name)
    else:
        EXCEL_LOG.error('The file format is not supported!')
        pass


if __name__ == '__main__':
    data1 = read_xls_column('./tencent.xls')
